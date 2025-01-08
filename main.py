from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import random
import re


def load_workbook_and_sheet(file_path: str, sheet_name: str) -> tuple[Workbook, Worksheet]:
    try:
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        sheet = workbook[sheet_name]
        return workbook, sheet
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        exit()
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        exit()


def initialize_trie(sheet):
    from TrieNode import Trie
    trie = Trie()
    all_pokemon_names = []

    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
        pokemon_name = row[1]
        if pokemon_name:  # Ensure name is not None
            pokemon_name = pokemon_name.strip().lower()
            all_pokemon_names.append(pokemon_name)
            trie.insert(pokemon_name)

    print("\n")
    print("Trie has been initialized...")
    print("\n")
    return all_pokemon_names, trie


def suggest_pokemon(all_pokemon: list) -> None:
    total_pokemon = len(all_pokemon)
    results = []
    for i in range(3):
        results.append(all_pokemon[random.randint(0, total_pokemon)])
    print(f"Suggestions and Formatting: {results[0].title()} | {results[1].title()} | {results[2].title()}")


def check_pokemon_input(pokemon_inquiry: str) -> bool:
    # Check if input contains only letters, spaces, and parentheses
    if not re.match(r'^[a-zA-Z\s()]+$', pokemon_inquiry):
        print("Invalid input. Please enter a valid Pokémon name.")
        return False

    print(f"Valid input: {pokemon_inquiry}")
    return True


def find_pokemon(sheet, inquiry_column: int, pokemon_name: str):
    pokemon_name_column = inquiry_column - 1  # Convert to zero-based index
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
        if row[pokemon_name_column] and row[pokemon_name_column].strip().lower() == pokemon_name.lower():
            header_row = [cell.value for cell in sheet[1]]
            min_length = min(len(header_row), len(row))
            pokemon_dict = {header_row[i]: row[i] for i in range(min_length)}
            return pokemon_dict

    print(f'Pokemon "{pokemon_name}" is not found in Pokemon Name List | Column: {inquiry_column}')
    return None


def print_pokemon_data(pokemon_data: dict) -> None:
    if pokemon_data:
        print("Pokemon Data:")
        for key, value in pokemon_data.items():
            if value and key != "Total":
                print(f'{key}: {value}')


def find_similar_spelling(trie, pokemon_name: str) -> list[str]:
    # Use Trie functionality to find words with similar prefixes
    # Need to add more in depth querying params
    prefix = pokemon_name[:3].lower()  # Example: Use first 2 characters for prefix search
    return trie.get_words_with_prefix(prefix)


def query_trie(trie, pokemon_name: str):
    suggestions = find_similar_spelling(trie, pokemon_name)
    if suggestions:
        print(f"Did you mean: {', '.join(suggestions)}?")
    else:
        print("No similar Pokémon names found.")


def main():
    active_wb = "excel_wb/National_Pokedex_Gen_9.xlsx"
    sheet_name = "Pokedex"
    inquiry_column = 2

    workbook, sheet = load_workbook_and_sheet(active_wb, sheet_name)
    all_poke_names, trie = initialize_trie(sheet)

    while True:
        suggest_pokemon(all_poke_names)
        pokemon_inquiry = input("Enter Pokémon name(enter 'q1' to exit): ").strip()
        if pokemon_inquiry.lower() == "q1":
            print("Exiting....")
            break

        if not check_pokemon_input(pokemon_inquiry):
            print("\n")
            continue

        pokemon_data = find_pokemon(sheet, inquiry_column, pokemon_inquiry)
        if pokemon_data:
            print_pokemon_data(pokemon_data)
        else:
            query_trie(trie, pokemon_inquiry)
        print("_____________________________\n")


if __name__ == "__main__":
    main()
